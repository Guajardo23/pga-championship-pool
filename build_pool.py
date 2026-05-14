"""Build PGA Championship Pool spreadsheet."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# --- Colors ---
NAVY = "00205B"
DARK_NAVY = "001740"
GOLD = "C8A951"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MED_GRAY = "D9D9D9"
MONEY_FMT = '$#,##0'

header_font = Font(name="Calibri", bold=True, color=WHITE, size=12)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
title_font = Font(name="Calibri", bold=True, color=DARK_NAVY, size=16)
money_font = Font(name="Calibri", size=11)
thin_border = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def style_range(ws, start_row, end_row, cols, alt=True):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if alt and (r - start_row) % 2 == 1:
                cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

# ============================================================
# SHEET 1: Golfer Earnings
# ============================================================
ws1 = wb.active
ws1.title = "Golfer Earnings"
ws1.sheet_properties.tabColor = NAVY

ws1.cell(row=1, column=1, value="2026 PGA Championship Pool - Golfer Earnings").font = title_font
ws1.merge_cells("A1:B1")

ws1["A3"] = "Golfer"
ws1["B3"] = "Earnings"
style_header(ws1, 3, 2)

# PGA Championship 2026 field (from ESPN API / pgatour.com)
golfers = [
    "Scottie Scheffler", "Rory McIlroy", "Xander Schauffele", "Jon Rahm",
    "Collin Morikawa", "Ludvig \u00c5berg", "Hideki Matsuyama", "Brooks Koepka",
    "Tommy Fleetwood", "Matt Fitzpatrick", "Cameron Young", "Bryson DeChambeau",
    "Viktor Hovland", "Patrick Cantlay", "Tyrrell Hatton", "Shane Lowry",
    "Sahith Theegala", "Wyndham Clark", "Sam Burns", "Robert MacIntyre",
    "Rickie Fowler", "Min Woo Lee", "Sungjae Im", "Keegan Bradley",
    "Justin Thomas", "Jordan Spieth", "Adam Scott", "Sepp Straka",
    "Russell Henley", "Tom McKibbin", "Nicolai H\u00f8jgaard", "Rasmus H\u00f8jgaard",
    "Corey Conners", "Maverick McNealy", "Alex Fitzpatrick", "Jacob Bridgeman",
    "Chris Gotterup", "Max Greyserman", "Jason Day", "Cameron Smith",
    "Harris English", "Brian Harman", "Daniel Berger", "Justin Rose",
    "Aaron Rai", "Patrick Reed", "Nick Taylor", "Alex Noren",
    "Si Woo Kim", "Thomas Detry", "Akshay Bhatia", "Kurt Kitayama",
    "Davis Riley", "Max Homa", "Billy Horschel", "Matt McCarty",
    "Joaqu\u00edn Niemann", "Christiaan Bezuidenhout", "Keith Mitchell",
    "Andrew Novak", "Emiliano Grillo", "Garrick Higgo", "Chris Kirk",
    "Tom Hoge", "Michael Kim", "Ben Griffin", "Denny McCarthy",
    "Dustin Johnson", "Ryan Fox", "Gary Woodland", "Lucas Glover",
    "J.J. Spaun", "Andrew Putnam", "Bud Cauley", "Marco Penge",
    "Aldrich Potgieter", "Sam Stevens", "Michael Brennan", "Brian Campbell",
    "Haotong Li", "Sami V\u00e4lim\u00e4ki", "David Lipsky", "Taylor Pendrith",
    "Rico Hoey", "Casey Jarvis", "Ryo Hisatsune", "Harry Hall",
    "Joe Highsmith", "Stephan Jaeger", "Michael Thorbjornsen",
    "Pierceson Coody", "J.T. Poston", "Adam Schenk", "Matt Wallace",
    "Kristoffer Reitan", "Rasmus Neergaard-Petersen", "Ryan Gerard",
    "Braden Shattuck", "David Puig", "Michael Block",
]

for i, name in enumerate(golfers):
    row = 4 + i
    ws1.cell(row=row, column=1, value=name)
    ws1.cell(row=row, column=2, value=0).number_format = MONEY_FMT

# Leave 20 blank rows for additional golfers
total_golfer_rows = len(golfers) + 20
for i in range(len(golfers), total_golfer_rows):
    row = 4 + i
    ws1.cell(row=row, column=2).number_format = MONEY_FMT

last_golfer_row = 4 + total_golfer_rows - 1
style_range(ws1, 4, last_golfer_row, 2)

ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 16

# ============================================================
# SHEET 2: Pool Entries
# ============================================================
ws2 = wb.create_sheet("Pool Entries")
ws2.sheet_properties.tabColor = NAVY

ws2.cell(row=1, column=1, value="2026 PGA Championship Pool - Entries & Standings").font = title_font
ws2.merge_cells("A1:H1")

headers = ["Contestant", "Pick 1", "Pick 2", "Pick 3", "Pick 4", "Pick 5",
           "Total Earnings", "Rank"]
for c, h in enumerate(headers, 1):
    ws2.cell(row=3, column=c, value=h)
style_header(ws2, 3, len(headers))

# 30 contestant rows with VLOOKUP formulas
MAX_CONTESTANTS = 30
lookup_range = f"'Golfer Earnings'!$A$4:$B${last_golfer_row}"

for r in range(4, 4 + MAX_CONTESTANTS):
    for pick_col in range(2, 7):
        ws2.cell(row=r, column=pick_col).number_format = '@'

    vlookups = []
    for pick_col_letter in ["B", "C", "D", "E", "F"]:
        vlookups.append(
            f'IF({pick_col_letter}{r}="",0,IFERROR(VLOOKUP({pick_col_letter}{r},{lookup_range},2,FALSE),0))'
        )
    formula = "=" + "+".join(vlookups)
    cell = ws2.cell(row=r, column=7, value=formula)
    cell.number_format = MONEY_FMT
    cell.font = Font(name="Calibri", bold=True, size=11)

    ws2.cell(row=r, column=8,
             value=f'=IF(A{r}="","",RANK(G{r},G$4:G${3+MAX_CONTESTANTS},0))').font = money_font

# Pre-populate contestant entries
entries = [
    ("JJ",      ["Rory McIlroy", "Scottie Scheffler", "Xander Schauffele", "Tommy Fleetwood", "Matt Fitzpatrick"]),
    ("Geoff",   ["Rory McIlroy", "Scottie Scheffler", "Jon Rahm", "Brooks Koepka", "Cameron Young"]),
    ("Brian",   ["Scottie Scheffler", "Rory McIlroy", "Cameron Young", "Collin Morikawa", "Tyrrell Hatton"]),
    ("Brandon", ["Rory McIlroy", "Scottie Scheffler", "Matt Fitzpatrick", "Cameron Young", "Ludvig \u00c5berg"]),
    ("Ben",     ["Scottie Scheffler", "Rory McIlroy", "Cameron Young", "Jacob Bridgeman", "Alex Fitzpatrick"]),
    ("Milo",    ["Cameron Young", "Matt Fitzpatrick", "Scottie Scheffler", "Rickie Fowler", "Brooks Koepka"]),
    ("Billy",   ["Rory McIlroy", "Scottie Scheffler", "Cameron Young", "Matt Fitzpatrick", "Ludvig \u00c5berg"]),
    ("Beto",    ["Scottie Scheffler", "Rory McIlroy", "Cameron Young", "Matt Fitzpatrick", "Justin Rose"]),
    ("Benny",   ["Collin Morikawa", "Xander Schauffele", "Brooks Koepka", "Rory McIlroy", "Scottie Scheffler"]),
    ("Paul",    ["Aaron Rai", "Max McGreevy", "Chris Gotterup", "Jacob Bridgeman", "Sam Burns"]),
    ("Nick",    ["Scottie Scheffler", "Rory McIlroy", "Cameron Young", "Matt Fitzpatrick", "Rickie Fowler"]),
    ("Pat",     ["Justin Rose", "Keegan Bradley", "Tommy Fleetwood", "Cameron Young", "Sepp Straka"]),
]
for i, (name, picks) in enumerate(entries):
    row = 4 + i
    ws2.cell(row=row, column=1, value=name)
    for j, pick in enumerate(picks):
        ws2.cell(row=row, column=2 + j, value=pick)

last_entry_row = 3 + MAX_CONTESTANTS
style_range(ws2, 4, last_entry_row, len(headers))

ws2.column_dimensions["A"].width = 22
for col in ["B", "C", "D", "E", "F"]:
    ws2.column_dimensions[col].width = 22
ws2.column_dimensions["G"].width = 18
ws2.column_dimensions["H"].width = 8

# ============================================================
# SHEET 3: Instructions
# ============================================================
ws3 = wb.create_sheet("Instructions")
ws3.sheet_properties.tabColor = "333333"

instructions = [
    ("How to Use This Spreadsheet", None),
    ("", None),
    ("Step 1: Enter Contestants", None),
    ("Go to the 'Pool Entries' tab. Type each person's name in column A, then their 5 golfer picks in columns B-F.", None),
    ("Golfer names must match exactly what's in the 'Golfer Earnings' tab.", None),
    ("", None),
    ("Step 2: After the Tournament", None),
    ("Go to the 'Golfer Earnings' tab. Enter each golfer's final earnings in column B.", None),
    ("The 'Pool Entries' tab will auto-calculate totals and rankings.", None),
    ("", None),
    ("Notes:", None),
    ("- If a pick doesn't match a golfer name, it counts as $0 (no error).", None),
    ("- Missed cut / WD golfers: just leave their earnings at $0.", None),
    ("- You can add more golfers beyond the pre-populated list.", None),
    ("- Rank column auto-sorts by highest total earnings.", None),
]

for i, (text, _) in enumerate(instructions):
    cell = ws3.cell(row=i + 1, column=1, value=text)
    if i == 0:
        cell.font = Font(name="Calibri", bold=True, color=DARK_NAVY, size=14)
    elif text.startswith("Step") or text == "Notes:":
        cell.font = Font(name="Calibri", bold=True, size=11)
    else:
        cell.font = Font(name="Calibri", size=11)

ws3.column_dimensions["A"].width = 90

# Save
out = r"C:\Users\jjg\SandboxSpace\PGAPool\pga_pool_2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
